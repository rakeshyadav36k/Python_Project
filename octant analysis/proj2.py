import streamlit as st

# imported some libraries
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
import pandas as pd
import os
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from streamlit_option_menu import option_menu



start_time = datetime.now()
st.set_page_config(page_title="Project 2",page_icon=":tada:",layout="wide")

with st.sidebar:
    select = option_menu("Main Menu",["Home","Browse File","Add path"],icons=['house','files','folder'],menu_icon="laptop",default_index=0)

if select =="Home":
    st.title("CS384 - PROJECT 2")
    st.subheader('''Built by : 
    Navneet Kumar Chouhan (2001CB37) 
    Rakesh Kumar Yadav (2001CB43) ''')

        
if select=="Add path":
    st.subheader("Add Path")
    folderpath = st.text_input('Enter folder path:')
    mod_input = st.text_input('Enter the Mod value')
    if st.button('Compute'):
        filenames = []
        path_out = 'output'
        cwd = os.getcwd()                   # this code is used for taking multiple input file
        os.chdir(folderpath)
        for file in os.listdir():
            file_namei = os.path.basename(file)
            file_name = os.path.splitext(file_namei)[0]
            filenames.append(file_name)
            file_ext = os.path.splitext(file_namei)[1]
            
            if file_ext=='.xlsx':
                wb = load_workbook(file)     # I have opened the Excel file
                sheet = wb.worksheets[0]

                df = pd.read_excel(file)     # df is a data frame in which we put the data of excel file using pandas


                top = Side(border_style='thin',color="000000")            # this whole code is for the border of a shell
                bottom = Side(border_style='thin',color="000000")
                left = Side(border_style='thin',color="000000")
                right = Side(border_style='thin',color="000000")

                border = Border(top=top,bottom=bottom,left=left,right=right)


                pattern = PatternFill(patternType='solid', fgColor="FFFF00")   # this code is for yellow color


                sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
                sheet.cell(row=1, column=6).value = "V Avg" 
                sheet.cell(row=1, column=7).value = "W Avg"  
                sheet.cell(row=1, column=8).value = "U'=U - U avg" 
                sheet.cell(row=1, column=9).value = "V'=V - V avg" 
                sheet.cell(row=1, column=10).value = "W'=W - W avg"

                Uavg = df['U'].mean()     # calculated the mean using pandas
                Vavg = df['V'].mean()
                Wavg = df['W'].mean()

                sheet.cell(row=2, column=5).value = Uavg          # added the average values in the sheet
                sheet.cell(row=2, column=6).value = Vavg          # not converted these values in upto three decimal because these valuse are very less and on keep it in upto three decimal it showing zero.
                sheet.cell(row=2, column=7).value = Wavg

                l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
                l2 = df['V']
                l3 = df['W']


                # creating three lists l4,l5 & l6 which contains the values of U', V' & W'

                #********************   
                l4=[]
                for i in l1:
                    a = i - Uavg
                    l4.append(a)

                for i in range(2,len(l1)+2):
                    sheet.cell(row=i, column=8).value = format(l4[i-2],"0.3f")

                l5=[]
                for i in l2:
                    a = i - Vavg
                    l5.append(a)

                for i in range(2,len(l2)+2):
                    sheet.cell(row=i, column=9).value = format(l5[i-2],"0.3f")

                l6=[]
                for i in l3:
                    a = i - Wavg
                    l6.append(a)

                for i in range(2,len(l3)+2):
                    sheet.cell(row=i, column=10).value = format(l6[i-2],"0.3f")

                #************************************

                # here we have created a header "Octant" and print the values of octants in excel file

                sheet.cell(row=1, column=11).value = "Octant"



                for i in range(0,len(l1)):
                    if(l4[i]>0 and l5[i]>0):
                        if(l6[i]>0):
                            sheet.cell(row=i+2, column=11).value = "+1"
                        else:                                               # this tells whether the octant is +1 or -1
                            sheet.cell(row=i+2, column=11).value = "-1"
                    elif(l4[i]<0 and l5[i]>0):
                        if(l6[i]>0):
                            sheet.cell(row=i+2, column=11).value = "+2"
                        else:                                               # this tells whether the octant is +2 or -2
                            sheet.cell(row=i+2, column=11).value = "-2"
                    elif(l4[i]<0 and l5[i]<0):
                        if(l6[i]>0):
                            sheet.cell(row=i+2, column=11).value = "+3"
                        else:                                                # this tells whether the octant is +3 or -3
                            sheet.cell(row=i+2, column=11).value = "-3"
                    elif(l4[i]>0 and l5[i]<0):
                        if(l6[i]>0):
                            sheet.cell(row=i+2, column=11).value = "+4"
                        else:                                                 # this tells whether the octant is +4 or -4
                            sheet.cell(row=i+2, column=11).value = "-4"


                # this list l7 contains the all octants values
                l7=[]
                for i in range(len(l1)):
                    x=sheet.cell(row=i+2,column=11).value
                    l7.append(int(x))


                sheet['N1']="Overall Octant Count"  # this is basicallly printed the header
                sheet['N3']="Octant ID"  
                sheet['N3'].border = border  
                sheet['N4']="Overall count"  
                sheet['N4'].border = border  
                sheet['O3'] = "+1"
                sheet['O3'].border = border
                sheet['P3'] = "-1"
                sheet['P3'].border = border
                sheet['Q3'] = "+2"
                sheet['Q3'].border = border
                sheet['R3'] = "-2"
                sheet['R3'].border = border
                sheet['S3'] = "+3"
                sheet['S3'].border = border
                sheet['T3'] = "-3"
                sheet['T3'].border = border
                sheet['U3'] = "+4"
                sheet['U3'].border = border
                sheet['V3'] = "-4"
                sheet['V3'].border = border

                ctpos1 = ctneg1 = ctpos2 = ctneg2 = ctpos3 = ctneg3 = ctpos4 = ctneg4 = 0  # these variables are total no each octant present

                for i in range(0,len(l1)):
                    if(l4[i]>0 and l5[i]>0):
                        if(l6[i]>0):
                            ctpos1 += 1               # total count of octant no +1 & -1
                        else:
                            ctneg1 += 1
                    elif(l4[i]<0 and l5[i]>0):
                        if(l6[i]>0):
                            ctpos2 += 1                 # total count of octant no +2 & -2
                        else:
                            ctneg2 += 1
                    elif(l4[i]<0 and l5[i]<0):
                        if(l6[i]>0):
                            ctpos3 += 1                 # total count of octant no +3 & -3
                        else:
                            ctneg3 += 1
                    elif(l4[i]>0 and l5[i]<0):
                        if(l6[i]>0):
                            ctpos4 += 1                 # total count of octant no +4 & -4
                        else:
                            ctneg4 += 1


                #  we have inserted the values of total no of each count

                sheet.cell(row=4, column=15).value = ctpos1
                sheet.cell(row=4, column=15).border = border
                sheet.cell(row=4, column=16).value = ctneg1
                sheet.cell(row=4, column=16).border = border
                sheet.cell(row=4, column=17).value = ctpos2
                sheet.cell(row=4, column=17).border = border
                sheet.cell(row=4, column=18).value = ctneg2
                sheet.cell(row=4, column=18).border = border
                sheet.cell(row=4, column=19).value = ctpos3
                sheet.cell(row=4, column=19).border = border
                sheet.cell(row=4, column=20).value = ctneg3
                sheet.cell(row=4, column=20).border = border
                sheet.cell(row=4, column=21).value = ctpos4
                sheet.cell(row=4, column=21).border = border
                sheet.cell(row=4, column=22).value = ctneg4
                sheet.cell(row=4, column=22).border = border


                mod = int(mod_input)    # this is a user defined mod value
                if(len(l7)%mod!=0):           
                    p = len(l7)//mod + 1           #variable p is no of partitions
                else:
                    p = len(l7)//mod 

                sheet['M4'] = "Mod" + " " +str(mod)

                A = []              # here we have taken a list A which contains another list B(list B contains the octants values of partition)
                x=0
                new_mod = mod
                for i in range(p):
                    B = []
                    for j in range(x,x + new_mod):
                        B.append(l7[j])
                    x+=mod
                    if((len(l7)-x)<mod):
                            new_mod = len(l7) - x
                    A.append(B)


                new_octant = [1,-1,2,-2,3,-3,4,-4]   #this is a list of octant values
                for i in range(p):
                    if(mod*(i+1)<len(l7)):
                        sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(mod*(i+1)-1)
                        sheet.cell(row=i+5, column=14).border = border
                    else:
                        sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(len(l7)-1)
                        sheet.cell(row=i+5, column=14).border = border
                    for j in range(8):
                        sheet.cell(row=i+5, column=15+j).value = A[i].count(new_octant[j])
                        sheet.cell(row=i+5, column=15+j).border = border          #we have counted the octant values in a range mod value

                ################################
                sheet.cell(row= 1,  column=35).value = "Overall Transition Count"    # it is simply written the headers
                sheet.cell(row=3, column=35).value = "Octant #"
                sheet.cell(row=3, column=35).border = border

                octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
                for i in range(8):
                    sheet.cell(row=4+i, column=35).value = octant[i]
                    sheet.cell(row=3, column=36+i).value = octant[i]
                    sheet.cell(row=4+i, column=35).border = border
                    sheet.cell(row=3, column=36+i).border = border

                sheet.cell(row=4, column=34).value = "From"
                sheet.cell(row=2, column=36).value = "To"

                # now are defining the transition from one octants to another octant using 64 variables 

                r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
                r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
                r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
                r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
                r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
                r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
                r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
                r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0 

                for i in range(len(l7)-1):      # these codes are to count the each transition values
                    if(l7[i]==1):
                        if(l7[i+1]==1):
                            r1c1+=1
                        elif(l7[i+1]==-1):
                            r1c2+=1
                        elif(l7[i+1]==2):
                            r1c3+=1
                        elif(l7[i+1]==-2):
                            r1c4+=1
                        elif(l7[i+1]==3):
                            r1c5+=1
                        elif(l7[i+1]==-3):
                            r1c6+=1
                        elif(l7[i+1]==4):
                            r1c7+=1
                        elif(l7[i+1]==-4):
                            r1c8+=1

                    if(l7[i]==-1):
                        if(l7[i+1]==1):
                            r2c1+=1
                        elif(l7[i+1]==-1):
                            r2c2+=1
                        elif(l7[i+1]==2):
                            r2c3+=1
                        elif(l7[i+1]==-2):
                            r2c4+=1
                        elif(l7[i+1]==3):
                            r2c5+=1
                        elif(l7[i+1]==-3):
                            r2c6+=1
                        elif(l7[i+1]==4):
                            r2c7+=1
                        elif(l7[i+1]==-4):
                            r2c8+=1

                    if(l7[i]==2):
                        if(l7[i+1]==1):
                            r3c1+=1
                        elif(l7[i+1]==-1):
                            r3c2+=1
                        elif(l7[i+1]==2):
                            r3c3+=1
                        elif(l7[i+1]==-2):
                            r3c4+=1
                        elif(l7[i+1]==3):
                            r3c5+=1
                        elif(l7[i+1]==-3):
                            r3c6+=1
                        elif(l7[i+1]==4):
                            r3c7+=1
                        elif(l7[i+1]==-4):
                            r3c8+=1

                    if(l7[i]==-2):
                        if(l7[i+1]==1):
                            r4c1+=1
                        elif(l7[i+1]==-1):
                            r4c2+=1
                        elif(l7[i+1]==2):
                            r4c3+=1
                        elif(l7[i+1]==-2):
                            r4c4+=1
                        elif(l7[i+1]==3):
                            r4c5+=1
                        elif(l7[i+1]==-3):
                            r4c6+=1
                        elif(l7[i+1]==4):
                            r4c7+=1
                        elif(l7[i+1]==-4):
                            r4c8+=1

                    if(l7[i]==3):
                        if(l7[i+1]==1):
                            r5c1+=1
                        elif(l7[i+1]==-1):
                            r5c2+=1
                        elif(l7[i+1]==2):
                            r5c3+=1
                        elif(l7[i+1]==-2):
                            r5c4+=1
                        elif(l7[i+1]==3):
                            r5c5+=1
                        elif(l7[i+1]==-3):
                            r5c6+=1
                        elif(l7[i+1]==4):
                            r5c7+=1
                        elif(l7[i+1]==-4):
                            r5c8+=1

                    if(l7[i]==-3):
                        if(l7[i+1]==1):
                            r6c1+=1
                        elif(l7[i+1]==-1):
                            r6c2+=1
                        elif(l7[i+1]==2):
                            r6c3+=1
                        elif(l7[i+1]==-2):
                            r6c4+=1
                        elif(l7[i+1]==3):
                            r6c5+=1
                        elif(l7[i+1]==-3):
                            r6c6+=1
                        elif(l7[i+1]==4):
                            r6c7+=1
                        elif(l7[i+1]==-4):
                            r6c8+=1

                    if(l7[i]==4):
                        if(l7[i+1]==1):
                            r7c1+=1
                        elif(l7[i+1]==-1):
                            r7c2+=1
                        elif(l7[i+1]==2):
                            r7c3+=1
                        elif(l7[i+1]==-2):
                            r7c4+=1
                        elif(l7[i+1]==3):
                            r7c5+=1
                        elif(l7[i+1]==-3):
                            r7c6+=1
                        elif(l7[i+1]==4):
                            r7c7+=1
                        elif(l7[i+1]==-4):
                            r7c8+=1

                    if(l7[i]==-4):
                        if(l7[i+1]==1):
                            r8c1+=1
                        elif(l7[i+1]==-1):
                            r8c2+=1
                        elif(l7[i+1]==2):
                            r8c3+=1
                        elif(l7[i+1]==-2):
                            r8c4+=1
                        elif(l7[i+1]==3):
                            r8c5+=1
                        elif(l7[i+1]==-3):
                            r8c6+=1
                        elif(l7[i+1]==4):
                            r8c7+=1
                        elif(l7[i+1]==-4):
                            r8c8+=1

                    # here transition values are put in the list by which it will easily print
                l8 = [r1c1,r1c2,r1c3,r1c4,r1c5,r1c6,r1c7,r1c8]
                l9 = [r2c1,r2c2,r2c3,r2c4,r2c5,r2c6,r2c7,r2c8]
                l10 = [r3c1,r3c2,r3c3,r3c4,r3c5,r3c6,r3c7,r3c8]
                l11 = [r4c1,r4c2,r4c3,r4c4,r4c5,r4c6,r4c7,r4c8]
                l12 = [r5c1,r5c2,r5c3,r5c4,r5c5,r5c6,r5c7,r5c8]
                l13 = [r6c1,r6c2,r6c3,r6c4,r6c5,r6c6,r6c7,r6c8]
                l14 = [r7c1,r7c2,r7c3,r7c4,r7c5,r7c6,r7c7,r7c8]
                l15 = [r8c1,r8c2,r8c3,r8c4,r8c5,r8c6,r8c7,r8c8]

                for i in range(8):                    # here have printed the overall transition values
                    sheet.cell(row=4,column=36+i).value = l8[i]
                    sheet.cell(row=5,column=36+i).value = l9[i]
                    sheet.cell(row=6,column=36+i).value = l10[i]
                    sheet.cell(row=7,column=36+i).value = l11[i]
                    sheet.cell(row=8,column=36+i).value = l12[i]
                    sheet.cell(row=9,column=36+i).value = l13[i]
                    sheet.cell(row=10,column=36+i).value = l14[i]
                    sheet.cell(row=11,column=36+i).value = l15[i]

                for i in range(8):                    # here have bordered the overall transition values
                    sheet.cell(row=4,column=36+i).border = border
                    sheet.cell(row=5,column=36+i).border = border
                    sheet.cell(row=6,column=36+i).border = border
                    sheet.cell(row=7,column=36+i).border = border
                    sheet.cell(row=8,column=36+i).border = border
                    sheet.cell(row=9,column=36+i).border = border
                    sheet.cell(row=10,column=36+i).border = border
                    sheet.cell(row=11,column=36+i).border = border




                r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0        # we again defining the transition values are zero
                r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
                r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
                r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
                r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
                r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
                r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
                r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0

                octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
                y=0
                new_mod = mod
                for x in range(p):                  # these codes are to compute the transition values in the a particular interval
                    for i in range(y,y+new_mod-1):
                        if(l7[i]==1):
                            if(l7[i+1]==1):
                                r1c1+=1
                            elif(l7[i+1]==-1):
                                r1c2+=1
                            elif(l7[i+1]==2):
                                r1c3+=1
                            elif(l7[i+1]==-2):
                                r1c4+=1
                            elif(l7[i+1]==3):
                                r1c5+=1
                            elif(l7[i+1]==-3):
                                r1c6+=1
                            elif(l7[i+1]==4):
                                r1c7+=1
                            elif(l7[i+1]==-4):
                                r1c8+=1

                        if(l7[i]==-1):
                            if(l7[i+1]==1):
                                r2c1+=1
                            elif(l7[i+1]==-1):
                                r2c2+=1
                            elif(l7[i+1]==2):
                                r2c3+=1
                            elif(l7[i+1]==-2):
                                r2c4+=1
                            elif(l7[i+1]==3):
                                r2c5+=1
                            elif(l7[i+1]==-3):
                                r2c6+=1
                            elif(l7[i+1]==4):
                                r2c7+=1
                            elif(l7[i+1]==-4):
                                r2c8+=1

                        if(l7[i]==2):
                            if(l7[i+1]==1):
                                r3c1+=1
                            elif(l7[i+1]==-1):
                                r3c2+=1
                            elif(l7[i+1]==2):
                                r3c3+=1
                            elif(l7[i+1]==-2):
                                r3c4+=1
                            elif(l7[i+1]==3):
                                r3c5+=1
                            elif(l7[i+1]==-3):
                                r3c6+=1
                            elif(l7[i+1]==4):
                                r3c7+=1
                            elif(l7[i+1]==-4):
                                r3c8+=1

                        if(l7[i]==-2):
                            if(l7[i+1]==1):
                                r4c1+=1
                            elif(l7[i+1]==-1):
                                r4c2+=1
                            elif(l7[i+1]==2):
                                r4c3+=1
                            elif(l7[i+1]==-2):
                                r4c4+=1
                            elif(l7[i+1]==3):
                                r4c5+=1
                            elif(l7[i+1]==-3):
                                r4c6+=1
                            elif(l7[i+1]==4):
                                r4c7+=1
                            elif(l7[i+1]==-4):
                                r4c8+=1

                        if(l7[i]==3):
                            if(l7[i+1]==1):
                                r5c1+=1
                            elif(l7[i+1]==-1):
                                r5c2+=1
                            elif(l7[i+1]==2):
                                r5c3+=1
                            elif(l7[i+1]==-2):
                                r5c4+=1
                            elif(l7[i+1]==3):
                                r5c5+=1
                            elif(l7[i+1]==-3):
                                r5c6+=1
                            elif(l7[i+1]==4):
                                r5c7+=1
                            elif(l7[i+1]==-4):
                                r5c8+=1

                        if(l7[i]==-3):
                            if(l7[i+1]==1):
                                r6c1+=1
                            elif(l7[i+1]==-1):
                                r6c2+=1
                            elif(l7[i+1]==2):
                                r6c3+=1
                            elif(l7[i+1]==-2):
                                r6c4+=1
                            elif(l7[i+1]==3):
                                r6c5+=1
                            elif(l7[i+1]==-3):
                                r6c6+=1
                            elif(l7[i+1]==4):
                                r6c7+=1
                            elif(l7[i+1]==-4):
                                r6c8+=1

                        if(l7[i]==4):
                            if(l7[i+1]==1):
                                r7c1+=1
                            elif(l7[i+1]==-1):
                                r7c2+=1
                            elif(l7[i+1]==2):
                                r7c3+=1
                            elif(l7[i+1]==-2):
                                r7c4+=1
                            elif(l7[i+1]==3):
                                r7c5+=1
                            elif(l7[i+1]==-3):
                                r7c6+=1
                            elif(l7[i+1]==4):
                                r7c7+=1
                            elif(l7[i+1]==-4):
                                r7c8+=1

                        if(l7[i]==-4):
                            if(l7[i+1]==1):
                                r8c1+=1
                            elif(l7[i+1]==-1):
                                r8c2+=1
                            elif(l7[i+1]==2):
                                r8c3+=1
                            elif(l7[i+1]==-2):
                                r8c4+=1
                            elif(l7[i+1]==3):
                                r8c5+=1
                            elif(l7[i+1]==-3):
                                r8c6+=1
                            elif(l7[i+1]==4):
                                r8c7+=1
                            elif(l7[i+1]==-4):
                                r8c8+=1

                        sheet.cell(row=14+x*13,column=35).value = str(y)+"-"+str(y+new_mod-1)
                    y+=mod
                    if((len(l7)-y)<mod):
                        new_mod = len(l7) - y

                    #we are putting the transition values in a list of list
                    new_list = [[r1c1, r1c2, r1c3, r1c4, r1c5 , r1c6, r1c7, r1c8],[r2c1 , r2c2 , r2c3 , r2c4 , r2c5 , r2c6 , r2c7 , r2c8],[r3c1 , r3c2 , r3c3 , r3c4 , r3c5 , r3c6 , r3c7 , r3c8],[r4c1 , r4c2 , r4c3 , r4c4 , r4c5 , r4c6 , r4c7 , r4c8]
                    ,[r5c1 , r5c2 , r5c3 , r5c4 , r5c5 , r5c6 , r5c7 , r5c8],[r6c1 , r6c2 , r6c3 , r6c4 , r6c5 , r6c6 , r6c7 , r6c8],[r7c1 , r7c2 , r7c3 , r7c4 , r7c5 , r7c6 , r7c7 , r7c8],[ r8c1 , r8c2 , r8c3 , r8c4 , r8c5 , r8c6 , r8c7 , r8c8]]

                    r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
                    r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
                    r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
                    r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
                    r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
                    r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
                    r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
                    r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0

                    sheet.cell(row=13+x*13,column=35).value = "Mod Transition Count"
                    sheet.cell(row=14+x*13,column=36).value = "To" 
                    sheet.cell(row=16+x*13,column=34).value = "From"
                    sheet.cell(row=15+x*13,column=35).value = "Octant #"
                    sheet.cell(row=15+x*13,column=35).border = border

                    for j in range(8):
                        sheet.cell(row=16+x*13+j,column=35).value = octant[j]
                        sheet.cell(row=15+x*13,column=36+j).value = octant[j]
                        sheet.cell(row=16+x*13+j,column=35).border = border
                        sheet.cell(row=15+x*13,column=36+j).border = border

                        for k in range(8):     # here we have printed the mod transition values using loops
                            sheet.cell(row=16+x*13+j,column=36+k).value = new_list[j][k]
                            sheet.cell(row=16+x*13+j, column=36+k).border = border


                ##**************************************************************************************************************##


                sheet.cell(row=1,column=45).value = "Longest Sequence Length"
                sheet.cell(row=3,column=45).value = "Octant ##"
                sheet.cell(row=3,column=46).value = "Longest Sequence Length"
                sheet.cell(row=3,column=47).value = "Count"
                for i in range(3):
                    sheet.cell(row=3,column=45+i).border = border

                list_octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
                for i in range(8):
                    sheet.cell(row=4+i,column=45).value = list_octant[i]
                    sheet.cell(row=4+i,column=45).border = border

                octs=[]           # we put all the octants values in list octs
                for i in range(len(l1)):
                    x=sheet.cell(row=i+2,column=11).value
                    octs.append(int(x))


                list1=[]      #list containing length of subsequents
                count1=1
                for i in range(len(octs)-1):
                    if(octs[i]==1 and octs[i+1]==1):
                        count1+=1
                    elif(octs[i]==1 and octs[i+1]!=1):
                        list1.append(count1)
                        count1=1

                list_1=[]      #list containing length of subsequents
                count_1=1
                for i in range(len(octs)-1):
                    if(octs[i]==-1 and octs[i+1]==-1):
                        count_1+=1
                    elif(octs[i]==-1 and octs[i+1]!=-1):
                        list_1.append(count_1)
                        count_1=1

                list2=[]      #list containing length of subsequents
                count2=1
                for i in range(len(octs)-1):
                    if(octs[i]==2 and octs[i+1]==2):
                        count2+=1
                    elif(octs[i]==2 and octs[i+1]!=2):
                        list2.append(count2)
                        count2=1

                list_2=[]      #list containing length of subsequents
                count_2=1
                for i in range(len(octs)-1):
                    if(octs[i]==-2 and octs[i+1]==-2):
                        count_2+=1
                    elif(octs[i]==-2 and octs[i+1]!=-2):
                        list_2.append(count_2)
                        count_2=1

                list3=[]      #list containing length of subsequents
                count3=1
                for i in range(len(octs)-1):
                    if(octs[i]==3 and octs[i+1]==3):
                        count3+=1
                    elif(octs[i]==3 and octs[i+1]!=3):
                        list3.append(count3)
                        count3=1

                list_3=[]      #list containing length of subsequents
                count_3=1
                for i in range(len(octs)-1):
                    if(octs[i]==-3 and octs[i+1]==-3):
                        count_3+=1
                    elif(octs[i]==-3 and octs[i+1]!=-3):
                        list_3.append(count_3)
                        count_3=1

                list4=[]      #list containing length of subsequents
                count4=1
                for i in range(len(octs)-1):
                    if(octs[i]==4 and octs[i+1]==4):
                        count4+=1
                    elif(octs[i]==4 and octs[i+1]!=4):
                        list4.append(count4)
                        count4=1

                list_4=[]      #list containing length of subsequents
                count_4=1
                for i in range(len(octs)-1):
                    if(octs[i]==-4 and octs[i+1]==-4):
                        count_4+=1
                    elif(octs[i]==-4 and octs[i+1]!=-4):
                        list_4.append(count_4)
                        count_4=1

                # here we have printed the longest subsequence of octants
                sheet['AT4'] = max(list1)
                sheet['AT5'] = max(list_1)
                sheet['AT6'] = max(list2)
                sheet['AT7'] = max(list_2)
                sheet['AT8'] = max(list3)
                sheet['AT9'] = max(list_3)
                sheet['AT10'] = max(list4)
                sheet['AT11'] = max(list_4)

                # here we have printed the howmany times maximum subsequence have occur
                sheet['AU4'] = list1.count(max(list1))
                sheet['AU5'] = list_1.count(max(list_1))
                sheet['AU6'] = list2.count(max(list2))
                sheet['AU7'] = list_2.count(max(list_2))
                sheet['AU8'] = list3.count(max(list3))
                sheet['AU9'] = list_3.count(max(list_3))
                sheet['AU10'] = list4.count(max(list4))
                sheet['AU11'] = list_4.count(max(list_4))

                for i in range(4,12):
                    for j in range(46,48):
                        sheet.cell(row=i,column=j).border = border


                    #################################################################################################################
                #################################################################################################################

                for i in range(8):
                    sheet.cell(row=3,column=23+i).value = "rank of "+str(new_octant[i])
                    sheet.cell(row=3,column=23+i).border = border


                dictionary_ct={1:ctpos1,-1:ctneg1,2:ctpos2,-2:ctneg2,3:ctpos3,-3:ctneg3,4:ctpos4,-4:ctneg4}
                dictionary_ct=dict(sorted(dictionary_ct.items(), key=lambda item:item[1]))
                dictionary_ct=list(dictionary_ct.items())

                for i in range(8):                        # code to print rank 1 in overall count
                    if(dictionary_ct[i][0]==1):
                        sheet.cell(row=4,column=23).value=8-i
                        sheet.cell(row=4,column=23).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=23).fill = pattern
                    elif(dictionary_ct[i][0]==-1):
                        sheet.cell(row=4,column=24).value=8-i
                        sheet.cell(row=4,column=24).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=24).fill = pattern
                    elif(dictionary_ct[i][0]==2):
                        sheet.cell(row=4,column=25).value=8-i
                        sheet.cell(row=4,column=25).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=25).fill = pattern
                    elif(dictionary_ct[i][0]==-2):
                        sheet.cell(row=4,column=26).value=8-i
                        sheet.cell(row=4,column=26).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=26).fill = pattern
                    elif(dictionary_ct[i][0]==3):
                        sheet.cell(row=4,column=27).value=8-i
                        sheet.cell(row=4,column=27).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=27).fill = pattern
                    elif(dictionary_ct[i][0]==-3):
                        sheet.cell(row=4,column=28).value=8-i
                        sheet.cell(row=4,column=28).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=28).fill = pattern
                    elif(dictionary_ct[i][0]==4):
                        sheet.cell(row=4,column=29).value=8-i
                        sheet.cell(row=4,column=29).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=29).fill = pattern
                    elif(dictionary_ct[i][0]==-4):
                        sheet.cell(row=4,column=30).value=8-i
                        sheet.cell(row=4,column=30).border = border
                        if((8-i)==1):
                            sheet.cell(row=4, column=30).fill = pattern



                sheet['AE3'] = "Rank1 Octant ID"
                sheet['AE3'].border = border
                sheet['AF3'] = "Rank1 Octant Name"
                sheet['AF3'].border = border

                octant_id = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
                sheet['AE4']=dictionary_ct[7][0]
                sheet['AE4'].border = border
                sheet['AF4']=octant_id[str(dictionary_ct[7][0])]
                sheet['AF4'].border = border

                rank1=[]    # this list stores the number of rank 1 octants

                for i in range(p):   # made a dictionary to store the count value of each octants
                    mod_dictionary_ct={1:A[i].count(1), -1:A[i].count(-1),2:A[i].count(2),-2:A[i].count(-2),3:A[i].count(3),-3:A[i].count(-3),4:A[i].count(4),-4:A[i].count(-4)}
                    mod_dictionary_ct=dict(sorted(mod_dictionary_ct.items(),key=lambda item:item[1]))     # sorted the values in increasing order
                    mod_dictionary_ct=list(mod_dictionary_ct.items())    # and make a list of that dictionary

                    for j in range(8):                  # code to print the rank of octants in mod values
                        if(mod_dictionary_ct[j][0]==1):
                            sheet.cell(row=5+i,column=23).value=8-j
                            sheet.cell(row=5+i,column=23).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=23).fill = pattern
                        elif(mod_dictionary_ct[j][0]==-1):
                            sheet.cell(row=5+i,column=24).value=8-j
                            sheet.cell(row=5+i,column=24).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=24).fill = pattern
                        elif(mod_dictionary_ct[j][0]==2):
                            sheet.cell(row=5+i,column=25).value=8-j
                            sheet.cell(row=5+i,column=25).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=25).fill = pattern
                        elif(mod_dictionary_ct[j][0]==-2):
                            sheet.cell(row=5+i,column=26).value=8-j
                            sheet.cell(row=5+i,column=26).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=26).fill = pattern
                        elif(mod_dictionary_ct[j][0]==3):
                            sheet.cell(row=5+i,column=27).value=8-j
                            sheet.cell(row=5+i,column=27).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=27).fill = pattern
                        elif(mod_dictionary_ct[j][0]==-3):
                            sheet.cell(row=5+i,column=28).value=8-j
                            sheet.cell(row=5+i,column=28).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=28).fill = pattern
                        elif(mod_dictionary_ct[j][0]==4):
                            sheet.cell(row=5+i,column=29).value=8-j
                            sheet.cell(row=5+i,column=29).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=29).fill = pattern
                        elif(mod_dictionary_ct[j][0]==-4):
                            sheet.cell(row=5+i,column=30).value=8-j
                            sheet.cell(row=5+i,column=30).border = border
                            if((8-j)==1):
                                sheet.cell(row=5+i, column=30).fill = pattern

                    sheet.cell(row=5+i,column=31).value=mod_dictionary_ct[7][0]
                    sheet.cell(row=5+i,column=31).border = border
                    rank1.append(mod_dictionary_ct[7][0])                   
                    sheet.cell(row=5+i,column=32).value=octant_id[str(mod_dictionary_ct[7][0])]
                    sheet.cell(row=5+i,column=32).border = border



                # code to print the octant which 1 rank appear most

                sheet.cell(row=6+p,column=29).value="Octant ID"
                sheet.cell(row=6+p,column=29).border = border
                sheet.cell(row=6+p,column=30).value="Octant Name"
                sheet.cell(row=6+p,column=30).border= border
                sheet.cell(row=6+p,column=31).value="Count of Rank 1 Mod Values"
                sheet.cell(row=6+p,column=31).border=border

                for i in range(8):                           # code for rank 1 ID
                    sheet.cell(row=7+p+i,column=29).value=str(new_octant[i])
                    sheet.cell(row=7+p+i,column=29).border = border
                    sheet.cell(row=7+p+i,column=30).value=octant_id[str(new_octant[i])]
                    sheet.cell(row=7+p+i,column=30).border = border
                    sheet.cell(row=7+p+i,column=31).value=rank1.count(new_octant[i])
                    sheet.cell(row=7+p+i,column=31).border = border

                #################################################################################################################
                #################################################################################################################

                time = []
                for i in range(len(l1)):
                    x=sheet.cell(row=i+2,column=1).value
                    time.append(x)

                row_no1 = []     #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list1=[]      #list containing length of subsequents
                count1=1
                for i in range(len(octs)-1):
                    if(octs[i]==1 and octs[i+1]==1):
                        count1+=1
                    elif(octs[i]==1 and octs[i+1]!=1):
                        list1.append(count1)
                        row_no1.append(i-count1+1)
                        count1=1

                time_l1 = []        #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list1)-1):
                    if(list1[i]==max(list1)):
                        a=row_no1[i]
                        time_l1.append([time[a],time[a+max(list1)-1]])

                row_no_1 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list_1=[]      #list containing length of subsequents
                count_1=1
                for i in range(len(octs)-1):
                    if(octs[i]==-1 and octs[i+1]==-1):
                        count_1+=1
                    elif(octs[i]==-1 and octs[i+1]!=-1):
                        list_1.append(count_1)
                        row_no_1.append(i-count_1+1)
                        count_1=1

                time_l_1 = []           #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list_1)-1):
                    if(list_1[i]==max(list_1)):
                        a=row_no_1[i]
                        time_l_1.append([time[a],time[a+max(list_1)-1]])


                row_no2 = []            #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list2=[]      #list containing length of subsequents
                count2=1
                for i in range(len(octs)-1):
                    if(octs[i]==2 and octs[i+1]==2):
                        count2+=1
                    elif(octs[i]==2 and octs[i+1]!=2):
                        list2.append(count2)
                        row_no2.append(i-count2+1)
                        count2=1

                time_l2 = []          #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list2)-1):
                    if(list2[i]==max(list2)):
                        a=row_no2[i]
                        time_l2.append([time[a],time[a+max(list2)-1]])

                row_no_2 = []       #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list_2=[]      #list containing length of subsequents
                count_2=1
                for i in range(len(octs)-1):
                    if(octs[i]==-2 and octs[i+1]==-2):
                        count_2+=1
                    elif(octs[i]==-2 and octs[i+1]!=-2):
                        list_2.append(count_2)
                        row_no_2.append(i-count_2+1)
                        count_2=1

                time_l_2 = []         #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list_2)-1):
                    if(list_2[i]==max(list_2)):
                        a=row_no_2[i]
                        time_l_2.append([time[a],time[a+max(list_2)-1]])

                row_no3 = []           #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list3=[]      #list containing length of subsequents
                count3=1
                for i in range(len(octs)-1):
                    if(octs[i]==3 and octs[i+1]==3):
                        count3+=1
                    elif(octs[i]==3 and octs[i+1]!=3):
                        list3.append(count3)
                        row_no3.append(i-count3+1)
                        count3=1

                time_l3 = []           #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list3)-1):
                    if(list3[i]==max(list3)):
                        a=row_no3[i]
                        time_l3.append([time[a],time[a+max(list3)-1]])

                row_no_3 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list_3=[]      #list containing length of subsequents
                count_3=1
                for i in range(len(octs)-1):
                    if(octs[i]==-3 and octs[i+1]==-3):
                        count_3+=1
                    elif(octs[i]==-3 and octs[i+1]!=-3):
                        list_3.append(count_3)
                        row_no_3.append(i-count_3+1)
                        count_3=1

                time_l_3 = []             #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list_3)-1):
                    if(list_3[i]==max(list_3)):
                        a=row_no_3[i]
                        time_l_3.append([time[a],time[a+max(list_3)-1]])

                row_no4 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list4=[]      #list containing length of subsequents
                count4=1
                for i in range(len(octs)-1):
                    if(octs[i]==4 and octs[i+1]==4):
                        count4+=1
                    elif(octs[i]==4 and octs[i+1]!=4):
                        list4.append(count4)
                        row_no4.append(i-count4+1)
                        count4=1

                time_l4 = []            #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list4)-1):
                    if(list4[i]==max(list4)):
                        a=row_no4[i]
                        time_l4.append([time[a],time[a+max(list4)-1]])

                row_no_4 = []        #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
                list_4=[]      #list containing length of subsequents
                count_4=1
                for i in range(len(octs)-1):
                    if(octs[i]==-4 and octs[i+1]==-4):
                        count_4+=1
                    elif(octs[i]==-4 and octs[i+1]!=-4):
                        list_4.append(count_4)
                        row_no_4.append(i-count_4+1)
                        count_4=1

                time_l_4 = []              #in this list we are collecting the range of time of a subsequence.
                for i in range(len(list_4)-1):
                    if(list_4[i]==max(list_4)):
                        a=row_no_4[i]
                        time_l_4.append([time[a],time[a+max(list_4)-1]])


                sheet.cell(row=1,column=49).value = "Longest Subsquence Length with Range"
                sheet.cell(row=3,column=49).value = "Octant ##"
                sheet.cell(row=3,column=50).value = "Longest Subsequence Length"
                sheet.cell(row=3,column=51).value = "Count"

                octant=[1,-1,2,-2,3,-3,4,-4]
                i=4                                                  # the whole code is to print the time range of longest subsequence for octants
                for oct in octant:
                    sheet.cell(row=i+1,column=49).value="Time"
                    sheet.cell(row=i+1,column=50).value="From"
                    sheet.cell(row=i+1,column=51).value="To"
                    if oct==1:
                        sheet.cell(row=i,column=49).value="+1"
                        sheet.cell(row=i,column=50).value=max(list1)
                        sheet.cell(row=i,column=51).value=list1.count(max(list1))

                        l=0
                        while l<len(time_l1):
                            sheet.cell(row=i+2+l,column=50).value=time_l1[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l1[l][1]
                            l+=1

                    elif oct==-1:
                        sheet.cell(row=i,column=49).value="-1"
                        sheet.cell(row=i,column=50).value=max(list_1)
                        sheet.cell(row=i,column=51).value=list_1.count(max(list_1))

                        l=0
                        while l<len(time_l_1):
                            sheet.cell(row=i+2+l,column=50).value=time_l_1[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l_1[l][1]
                            l+=1

                    elif oct==2:
                        sheet.cell(row=i,column=49).value="+2"
                        sheet.cell(row=i,column=50).value=max(list2)
                        sheet.cell(row=i,column=51).value=list2.count(max(list2))

                        l=0
                        while l<len(time_l2):
                            sheet.cell(row=i+2+l,column=50).value=time_l2[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l2[l][1]
                            l+=1

                    elif oct==-2:
                        sheet.cell(row=i,column=49).value="-2"
                        sheet.cell(row=i,column=50).value=max(list_2)
                        sheet.cell(row=i,column=51).value=list_2.count(max(list_2))

                        l=0
                        while l<len(time_l_2):
                            sheet.cell(row=i+2+l,column=50).value=time_l_2[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l_2[l][1]
                            l+=1

                    elif oct==3:
                        sheet.cell(row=i,column=49).value="+3"
                        sheet.cell(row=i,column=50).value=max(list3)
                        sheet.cell(row=i,column=51).value=list3.count(max(list3))

                        l=0
                        while l<len(time_l3):
                            sheet.cell(row=i+2+l,column=50).value=time_l3[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l3[l][1]
                            l+=1

                    elif oct==-3:
                        sheet.cell(row=i,column=49).value="-3"
                        sheet.cell(row=i,column=50).value=max(list_3)
                        sheet.cell(row=i,column=51).value=list_3.count(max(list_3))

                        l=0
                        while l<len(time_l_3):
                            sheet.cell(row=i+2+l,column=50).value=time_l_3[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l_3[l][1]
                            l+=1

                    elif oct==4:
                        sheet.cell(row=i,column=49).value="+4"
                        sheet.cell(row=i,column=50).value=max(list4)
                        sheet.cell(row=i,column=51).value=list4.count(max(list4))

                        l=0
                        while l<len(time_l4):
                            sheet.cell(row=i+2+l,column=50).value=time_l4[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l4[l][1]
                            l+=1

                    elif oct==-4:
                        sheet.cell(row=i,column=49).value="-4"
                        sheet.cell(row=i,column=50).value=max(list_4)
                        sheet.cell(row=i,column=51).value=list_4.count(max(list_4))

                        l=0
                        while l<len(time_l_4):
                            sheet.cell(row=i+2+l,column=50).value=time_l_4[l][0]
                            sheet.cell(row=i+2+l,column=51).value=time_l_4[l][1]
                            l+=1

                    i+=l+2


                #####################################################

                for k in range(3,30):
                    for l in range(49,52):
                        sheet.cell(row=k, column=l).border = border



                for column in ["E","F","G","H","I","J","W","X","Y","Z","AB","AA","AC"]:
                    sheet.column_dimensions[column].width = 15                           # have set the width of the some column

                for column in ["AD","AE","AF","AI","AS","AT","AX"]:
                    sheet.column_dimensions[column].width = 30

                for column in ["O","P","Q","R","S","T","U","V","AJ","AK","AL","AM","AN","AO","AP","AQ","AU","AY"]:
                    sheet.column_dimensions[column].width = 6

                for column in ["AW"]:
                    sheet.column_dimensions[column].width = 38
                for column in ["N"]:
                    sheet.column_dimensions[column].width = 22

                os.chdir(cwd)                             # we save the output file and changing the directory to take a new input file
                os.chdir(path_out)
                time1=datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
                wb.save(str(file_name)+ '_mod' + '(' + str(mod)+ ')_' + str(time1) + '.xlsx')

                # os.chdir(cwd)
                os.chdir(folderpath)

        os.chdir(cwd)
        st.info("Completed !!")


######################################################
###                                                ###
###         Other part of code                     ###
###                                                ###
###                                                ###
######################################################


if select=="Browse File":
    st.header("Browse File")
    files = st.file_uploader("Upload a excel file",accept_multiple_files=True)    
    mod_input = st.text_input('Enter the Mod value')

    if st.button('Compute'):        
        df = pd.DataFrame
        for file in files:
            wb = load_workbook(file)
            sheet = wb.active


            df = pd.read_excel(file)     # df is a data frame in which we put the data of excel file using pandas


            top = Side(border_style='thin',color="000000")            # this whole code is for the border of a shell
            bottom = Side(border_style='thin',color="000000")
            left = Side(border_style='thin',color="000000")
            right = Side(border_style='thin',color="000000")

            border = Border(top=top,bottom=bottom,left=left,right=right)


            pattern = PatternFill(patternType='solid', fgColor="FFFF00")   # this code is for yellow color


            sheet.cell(row=1, column=5).value = "U Avg"       # Written the header 
            sheet.cell(row=1, column=6).value = "V Avg" 
            sheet.cell(row=1, column=7).value = "W Avg"  
            sheet.cell(row=1, column=8).value = "U'=U - U avg" 
            sheet.cell(row=1, column=9).value = "V'=V - V avg" 
            sheet.cell(row=1, column=10).value = "W'=W - W avg"

            Uavg = df['U'].mean()     # calculated the mean using pandas
            Vavg = df['V'].mean()
            Wavg = df['W'].mean()

            sheet.cell(row=2, column=5).value = Uavg          # added the average values in the sheet
            sheet.cell(row=2, column=6).value = Vavg          # not converted these values in upto three decimal because these valuse are very less and on keep it in upto three decimal it showing zero.
            sheet.cell(row=2, column=7).value = Wavg

            l1 = df['U']            # creating the list l1,l2,l3 which consist the element of U,V,W respectively
            l2 = df['V']
            l3 = df['W']


            # creating three lists l4,l5 & l6 which contains the values of U', V' & W'

            #********************   
            l4=[]
            for i in l1:
                a = i - Uavg
                l4.append(a)

            for i in range(2,len(l1)+2):
                sheet.cell(row=i, column=8).value = format(l4[i-2],"0.3f")

            l5=[]
            for i in l2:
                a = i - Vavg
                l5.append(a)

            for i in range(2,len(l2)+2):
                sheet.cell(row=i, column=9).value = format(l5[i-2],"0.3f")

            l6=[]
            for i in l3:
                a = i - Wavg
                l6.append(a)

            for i in range(2,len(l3)+2):
                sheet.cell(row=i, column=10).value = format(l6[i-2],"0.3f")

            #************************************

            # here we have created a header "Octant" and print the values of octants in excel file

            sheet.cell(row=1, column=11).value = "Octant"



            for i in range(0,len(l1)):
                if(l4[i]>0 and l5[i]>0):
                    if(l6[i]>0):
                        sheet.cell(row=i+2, column=11).value = "+1"
                    else:                                               # this tells whether the octant is +1 or -1
                        sheet.cell(row=i+2, column=11).value = "-1"
                elif(l4[i]<0 and l5[i]>0):
                    if(l6[i]>0):
                        sheet.cell(row=i+2, column=11).value = "+2"
                    else:                                               # this tells whether the octant is +2 or -2
                        sheet.cell(row=i+2, column=11).value = "-2"
                elif(l4[i]<0 and l5[i]<0):
                    if(l6[i]>0):
                        sheet.cell(row=i+2, column=11).value = "+3"
                    else:                                                # this tells whether the octant is +3 or -3
                        sheet.cell(row=i+2, column=11).value = "-3"
                elif(l4[i]>0 and l5[i]<0):
                    if(l6[i]>0):
                        sheet.cell(row=i+2, column=11).value = "+4"
                    else:                                                 # this tells whether the octant is +4 or -4
                        sheet.cell(row=i+2, column=11).value = "-4"


            # this list l7 contains the all octants values
            l7=[]
            for i in range(len(l1)):
                x=sheet.cell(row=i+2,column=11).value
                l7.append(int(x))


            sheet['N1']="Overall Octant Count"  # this is basicallly printed the header
            sheet['N3']="Octant ID"  
            sheet['N3'].border = border  
            sheet['N4']="Overall count"  
            sheet['N4'].border = border  
            sheet['O3'] = "+1"
            sheet['O3'].border = border
            sheet['P3'] = "-1"
            sheet['P3'].border = border
            sheet['Q3'] = "+2"
            sheet['Q3'].border = border
            sheet['R3'] = "-2"
            sheet['R3'].border = border
            sheet['S3'] = "+3"
            sheet['S3'].border = border
            sheet['T3'] = "-3"
            sheet['T3'].border = border
            sheet['U3'] = "+4"
            sheet['U3'].border = border
            sheet['V3'] = "-4"
            sheet['V3'].border = border

            ctpos1 = ctneg1 = ctpos2 = ctneg2 = ctpos3 = ctneg3 = ctpos4 = ctneg4 = 0  # these variables are total no each octant present

            for i in range(0,len(l1)):
                if(l4[i]>0 and l5[i]>0):
                    if(l6[i]>0):
                        ctpos1 += 1               # total count of octant no +1 & -1
                    else:
                        ctneg1 += 1
                elif(l4[i]<0 and l5[i]>0):
                    if(l6[i]>0):
                        ctpos2 += 1                 # total count of octant no +2 & -2
                    else:
                        ctneg2 += 1
                elif(l4[i]<0 and l5[i]<0):
                    if(l6[i]>0):
                        ctpos3 += 1                 # total count of octant no +3 & -3
                    else:
                        ctneg3 += 1
                elif(l4[i]>0 and l5[i]<0):
                    if(l6[i]>0):
                        ctpos4 += 1                 # total count of octant no +4 & -4
                    else:
                        ctneg4 += 1


            #  we have inserted the values of total no of each count

            sheet.cell(row=4, column=15).value = ctpos1
            sheet.cell(row=4, column=15).border = border
            sheet.cell(row=4, column=16).value = ctneg1
            sheet.cell(row=4, column=16).border = border
            sheet.cell(row=4, column=17).value = ctpos2
            sheet.cell(row=4, column=17).border = border
            sheet.cell(row=4, column=18).value = ctneg2
            sheet.cell(row=4, column=18).border = border
            sheet.cell(row=4, column=19).value = ctpos3
            sheet.cell(row=4, column=19).border = border
            sheet.cell(row=4, column=20).value = ctneg3
            sheet.cell(row=4, column=20).border = border
            sheet.cell(row=4, column=21).value = ctpos4
            sheet.cell(row=4, column=21).border = border
            sheet.cell(row=4, column=22).value = ctneg4
            sheet.cell(row=4, column=22).border = border


            mod = int(mod_input)    # this is a user defined mod value
            if(len(l7)%mod!=0):           
                p = len(l7)//mod + 1           #variable p is no of partitions
            else:
                p = len(l7)//mod 

            sheet['M4'] = "Mod" + " " +str(mod)

            A = []              # here we have taken a list A which contains another list B(list B contains the octants values of partition)
            x=0
            new_mod = mod
            for i in range(p):
                B = []
                for j in range(x,x + new_mod):
                    B.append(l7[j])
                x+=mod
                if((len(l7)-x)<mod):
                        new_mod = len(l7) - x
                A.append(B)


            new_octant = [1,-1,2,-2,3,-3,4,-4]   #this is a list of octant values
            for i in range(p):
                if(mod*(i+1)<len(l7)):
                    sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(mod*(i+1)-1)
                    sheet.cell(row=i+5, column=14).border = border
                else:
                    sheet.cell(row=i+5, column=14).value = str(mod*i)+"-"+str(len(l7)-1)
                    sheet.cell(row=i+5, column=14).border = border
                for j in range(8):
                    sheet.cell(row=i+5, column=15+j).value = A[i].count(new_octant[j])
                    sheet.cell(row=i+5, column=15+j).border = border          #we have counted the octant values in a range mod value

            ################################
            sheet.cell(row= 1,  column=35).value = "Overall Transition Count"    # it is simply written the headers
            sheet.cell(row=3, column=35).value = "Octant #"
            sheet.cell(row=3, column=35).border = border

            octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
            for i in range(8):
                sheet.cell(row=4+i, column=35).value = octant[i]
                sheet.cell(row=3, column=36+i).value = octant[i]
                sheet.cell(row=4+i, column=35).border = border
                sheet.cell(row=3, column=36+i).border = border

            sheet.cell(row=4, column=34).value = "From"
            sheet.cell(row=2, column=36).value = "To"

            # now are defining the transition from one octants to another octant using 64 variables 

            r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
            r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
            r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
            r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
            r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
            r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
            r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
            r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0 

            for i in range(len(l7)-1):      # these codes are to count the each transition values
                if(l7[i]==1):
                    if(l7[i+1]==1):
                        r1c1+=1
                    elif(l7[i+1]==-1):
                        r1c2+=1
                    elif(l7[i+1]==2):
                        r1c3+=1
                    elif(l7[i+1]==-2):
                        r1c4+=1
                    elif(l7[i+1]==3):
                        r1c5+=1
                    elif(l7[i+1]==-3):
                        r1c6+=1
                    elif(l7[i+1]==4):
                        r1c7+=1
                    elif(l7[i+1]==-4):
                        r1c8+=1

                if(l7[i]==-1):
                    if(l7[i+1]==1):
                        r2c1+=1
                    elif(l7[i+1]==-1):
                        r2c2+=1
                    elif(l7[i+1]==2):
                        r2c3+=1
                    elif(l7[i+1]==-2):
                        r2c4+=1
                    elif(l7[i+1]==3):
                        r2c5+=1
                    elif(l7[i+1]==-3):
                        r2c6+=1
                    elif(l7[i+1]==4):
                        r2c7+=1
                    elif(l7[i+1]==-4):
                        r2c8+=1

                if(l7[i]==2):
                    if(l7[i+1]==1):
                        r3c1+=1
                    elif(l7[i+1]==-1):
                        r3c2+=1
                    elif(l7[i+1]==2):
                        r3c3+=1
                    elif(l7[i+1]==-2):
                        r3c4+=1
                    elif(l7[i+1]==3):
                        r3c5+=1
                    elif(l7[i+1]==-3):
                        r3c6+=1
                    elif(l7[i+1]==4):
                        r3c7+=1
                    elif(l7[i+1]==-4):
                        r3c8+=1

                if(l7[i]==-2):
                    if(l7[i+1]==1):
                        r4c1+=1
                    elif(l7[i+1]==-1):
                        r4c2+=1
                    elif(l7[i+1]==2):
                        r4c3+=1
                    elif(l7[i+1]==-2):
                        r4c4+=1
                    elif(l7[i+1]==3):
                        r4c5+=1
                    elif(l7[i+1]==-3):
                        r4c6+=1
                    elif(l7[i+1]==4):
                        r4c7+=1
                    elif(l7[i+1]==-4):
                        r4c8+=1

                if(l7[i]==3):
                    if(l7[i+1]==1):
                        r5c1+=1
                    elif(l7[i+1]==-1):
                        r5c2+=1
                    elif(l7[i+1]==2):
                        r5c3+=1
                    elif(l7[i+1]==-2):
                        r5c4+=1
                    elif(l7[i+1]==3):
                        r5c5+=1
                    elif(l7[i+1]==-3):
                        r5c6+=1
                    elif(l7[i+1]==4):
                        r5c7+=1
                    elif(l7[i+1]==-4):
                        r5c8+=1

                if(l7[i]==-3):
                    if(l7[i+1]==1):
                        r6c1+=1
                    elif(l7[i+1]==-1):
                        r6c2+=1
                    elif(l7[i+1]==2):
                        r6c3+=1
                    elif(l7[i+1]==-2):
                        r6c4+=1
                    elif(l7[i+1]==3):
                        r6c5+=1
                    elif(l7[i+1]==-3):
                        r6c6+=1
                    elif(l7[i+1]==4):
                        r6c7+=1
                    elif(l7[i+1]==-4):
                        r6c8+=1

                if(l7[i]==4):
                    if(l7[i+1]==1):
                        r7c1+=1
                    elif(l7[i+1]==-1):
                        r7c2+=1
                    elif(l7[i+1]==2):
                        r7c3+=1
                    elif(l7[i+1]==-2):
                        r7c4+=1
                    elif(l7[i+1]==3):
                        r7c5+=1
                    elif(l7[i+1]==-3):
                        r7c6+=1
                    elif(l7[i+1]==4):
                        r7c7+=1
                    elif(l7[i+1]==-4):
                        r7c8+=1

                if(l7[i]==-4):
                    if(l7[i+1]==1):
                        r8c1+=1
                    elif(l7[i+1]==-1):
                        r8c2+=1
                    elif(l7[i+1]==2):
                        r8c3+=1
                    elif(l7[i+1]==-2):
                        r8c4+=1
                    elif(l7[i+1]==3):
                        r8c5+=1
                    elif(l7[i+1]==-3):
                        r8c6+=1
                    elif(l7[i+1]==4):
                        r8c7+=1
                    elif(l7[i+1]==-4):
                        r8c8+=1

                # here transition values are put in the list by which it will easily print
            l8 = [r1c1,r1c2,r1c3,r1c4,r1c5,r1c6,r1c7,r1c8]
            l9 = [r2c1,r2c2,r2c3,r2c4,r2c5,r2c6,r2c7,r2c8]
            l10 = [r3c1,r3c2,r3c3,r3c4,r3c5,r3c6,r3c7,r3c8]
            l11 = [r4c1,r4c2,r4c3,r4c4,r4c5,r4c6,r4c7,r4c8]
            l12 = [r5c1,r5c2,r5c3,r5c4,r5c5,r5c6,r5c7,r5c8]
            l13 = [r6c1,r6c2,r6c3,r6c4,r6c5,r6c6,r6c7,r6c8]
            l14 = [r7c1,r7c2,r7c3,r7c4,r7c5,r7c6,r7c7,r7c8]
            l15 = [r8c1,r8c2,r8c3,r8c4,r8c5,r8c6,r8c7,r8c8]

            for i in range(8):                    # here have printed the overall transition values
                sheet.cell(row=4,column=36+i).value = l8[i]
                sheet.cell(row=5,column=36+i).value = l9[i]
                sheet.cell(row=6,column=36+i).value = l10[i]
                sheet.cell(row=7,column=36+i).value = l11[i]
                sheet.cell(row=8,column=36+i).value = l12[i]
                sheet.cell(row=9,column=36+i).value = l13[i]
                sheet.cell(row=10,column=36+i).value = l14[i]
                sheet.cell(row=11,column=36+i).value = l15[i]

            for i in range(8):                    # here have bordered the overall transition values
                sheet.cell(row=4,column=36+i).border = border
                sheet.cell(row=5,column=36+i).border = border
                sheet.cell(row=6,column=36+i).border = border
                sheet.cell(row=7,column=36+i).border = border
                sheet.cell(row=8,column=36+i).border = border
                sheet.cell(row=9,column=36+i).border = border
                sheet.cell(row=10,column=36+i).border = border
                sheet.cell(row=11,column=36+i).border = border




            r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0        # we again defining the transition values are zero
            r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
            r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
            r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
            r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
            r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
            r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
            r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0

            octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
            y=0
            new_mod = mod
            for x in range(p):                  # these codes are to compute the transition values in the a particular interval
                for i in range(y,y+new_mod-1):
                    if(l7[i]==1):
                        if(l7[i+1]==1):
                            r1c1+=1
                        elif(l7[i+1]==-1):
                            r1c2+=1
                        elif(l7[i+1]==2):
                            r1c3+=1
                        elif(l7[i+1]==-2):
                            r1c4+=1
                        elif(l7[i+1]==3):
                            r1c5+=1
                        elif(l7[i+1]==-3):
                            r1c6+=1
                        elif(l7[i+1]==4):
                            r1c7+=1
                        elif(l7[i+1]==-4):
                            r1c8+=1

                    if(l7[i]==-1):
                        if(l7[i+1]==1):
                            r2c1+=1
                        elif(l7[i+1]==-1):
                            r2c2+=1
                        elif(l7[i+1]==2):
                            r2c3+=1
                        elif(l7[i+1]==-2):
                            r2c4+=1
                        elif(l7[i+1]==3):
                            r2c5+=1
                        elif(l7[i+1]==-3):
                            r2c6+=1
                        elif(l7[i+1]==4):
                            r2c7+=1
                        elif(l7[i+1]==-4):
                            r2c8+=1

                    if(l7[i]==2):
                        if(l7[i+1]==1):
                            r3c1+=1
                        elif(l7[i+1]==-1):
                            r3c2+=1
                        elif(l7[i+1]==2):
                            r3c3+=1
                        elif(l7[i+1]==-2):
                            r3c4+=1
                        elif(l7[i+1]==3):
                            r3c5+=1
                        elif(l7[i+1]==-3):
                            r3c6+=1
                        elif(l7[i+1]==4):
                            r3c7+=1
                        elif(l7[i+1]==-4):
                            r3c8+=1

                    if(l7[i]==-2):
                        if(l7[i+1]==1):
                            r4c1+=1
                        elif(l7[i+1]==-1):
                            r4c2+=1
                        elif(l7[i+1]==2):
                            r4c3+=1
                        elif(l7[i+1]==-2):
                            r4c4+=1
                        elif(l7[i+1]==3):
                            r4c5+=1
                        elif(l7[i+1]==-3):
                            r4c6+=1
                        elif(l7[i+1]==4):
                            r4c7+=1
                        elif(l7[i+1]==-4):
                            r4c8+=1

                    if(l7[i]==3):
                        if(l7[i+1]==1):
                            r5c1+=1
                        elif(l7[i+1]==-1):
                            r5c2+=1
                        elif(l7[i+1]==2):
                            r5c3+=1
                        elif(l7[i+1]==-2):
                            r5c4+=1
                        elif(l7[i+1]==3):
                            r5c5+=1
                        elif(l7[i+1]==-3):
                            r5c6+=1
                        elif(l7[i+1]==4):
                            r5c7+=1
                        elif(l7[i+1]==-4):
                            r5c8+=1

                    if(l7[i]==-3):
                        if(l7[i+1]==1):
                            r6c1+=1
                        elif(l7[i+1]==-1):
                            r6c2+=1
                        elif(l7[i+1]==2):
                            r6c3+=1
                        elif(l7[i+1]==-2):
                            r6c4+=1
                        elif(l7[i+1]==3):
                            r6c5+=1
                        elif(l7[i+1]==-3):
                            r6c6+=1
                        elif(l7[i+1]==4):
                            r6c7+=1
                        elif(l7[i+1]==-4):
                            r6c8+=1

                    if(l7[i]==4):
                        if(l7[i+1]==1):
                            r7c1+=1
                        elif(l7[i+1]==-1):
                            r7c2+=1
                        elif(l7[i+1]==2):
                            r7c3+=1
                        elif(l7[i+1]==-2):
                            r7c4+=1
                        elif(l7[i+1]==3):
                            r7c5+=1
                        elif(l7[i+1]==-3):
                            r7c6+=1
                        elif(l7[i+1]==4):
                            r7c7+=1
                        elif(l7[i+1]==-4):
                            r7c8+=1

                    if(l7[i]==-4):
                        if(l7[i+1]==1):
                            r8c1+=1
                        elif(l7[i+1]==-1):
                            r8c2+=1
                        elif(l7[i+1]==2):
                            r8c3+=1
                        elif(l7[i+1]==-2):
                            r8c4+=1
                        elif(l7[i+1]==3):
                            r8c5+=1
                        elif(l7[i+1]==-3):
                            r8c6+=1
                        elif(l7[i+1]==4):
                            r8c7+=1
                        elif(l7[i+1]==-4):
                            r8c8+=1

                    sheet.cell(row=14+x*13,column=35).value = str(y)+"-"+str(y+new_mod-1)
                y+=mod
                if((len(l7)-y)<mod):
                    new_mod = len(l7) - y

                #we are putting the transition values in a list of list
                new_list = [[r1c1, r1c2, r1c3, r1c4, r1c5 , r1c6, r1c7, r1c8],[r2c1 , r2c2 , r2c3 , r2c4 , r2c5 , r2c6 , r2c7 , r2c8],[r3c1 , r3c2 , r3c3 , r3c4 , r3c5 , r3c6 , r3c7 , r3c8],[r4c1 , r4c2 , r4c3 , r4c4 , r4c5 , r4c6 , r4c7 , r4c8]
                ,[r5c1 , r5c2 , r5c3 , r5c4 , r5c5 , r5c6 , r5c7 , r5c8],[r6c1 , r6c2 , r6c3 , r6c4 , r6c5 , r6c6 , r6c7 , r6c8],[r7c1 , r7c2 , r7c3 , r7c4 , r7c5 , r7c6 , r7c7 , r7c8],[ r8c1 , r8c2 , r8c3 , r8c4 , r8c5 , r8c6 , r8c7 , r8c8]]

                r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
                r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
                r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
                r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
                r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
                r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
                r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
                r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0

                sheet.cell(row=13+x*13,column=35).value = "Mod Transition Count"
                sheet.cell(row=14+x*13,column=36).value = "To" 
                sheet.cell(row=16+x*13,column=34).value = "From"
                sheet.cell(row=15+x*13,column=35).value = "Octant #"
                sheet.cell(row=15+x*13,column=35).border = border

                for j in range(8):
                    sheet.cell(row=16+x*13+j,column=35).value = octant[j]
                    sheet.cell(row=15+x*13,column=36+j).value = octant[j]
                    sheet.cell(row=16+x*13+j,column=35).border = border
                    sheet.cell(row=15+x*13,column=36+j).border = border

                    for k in range(8):     # here we have printed the mod transition values using loops
                        sheet.cell(row=16+x*13+j,column=36+k).value = new_list[j][k]
                        sheet.cell(row=16+x*13+j, column=36+k).border = border
                       


            ##**************************************************************************************************************##


            sheet.cell(row=1,column=45).value = "Longest Sequence Length"
            sheet.cell(row=3,column=45).value = "Octant ##"
            sheet.cell(row=3,column=46).value = "Longest Sequence Length"
            sheet.cell(row=3,column=47).value = "Count"
            for i in range(3):
                sheet.cell(row=3,column=45+i).border = border

            list_octant = ["+1","-1","+2","-2","+3","-3","+4","-4"]
            for i in range(8):
                sheet.cell(row=4+i,column=45).value = list_octant[i]
                sheet.cell(row=4+i,column=45).border = border

            octs=[]           # we put all the octants values in list octs
            for i in range(len(l1)):
                x=sheet.cell(row=i+2,column=11).value
                octs.append(int(x))


            list1=[]      #list containing length of subsequents
            count1=1
            for i in range(len(octs)-1):
                if(octs[i]==1 and octs[i+1]==1):
                    count1+=1
                elif(octs[i]==1 and octs[i+1]!=1):
                    list1.append(count1)
                    count1=1

            list_1=[]      #list containing length of subsequents
            count_1=1
            for i in range(len(octs)-1):
                if(octs[i]==-1 and octs[i+1]==-1):
                    count_1+=1
                elif(octs[i]==-1 and octs[i+1]!=-1):
                    list_1.append(count_1)
                    count_1=1

            list2=[]      #list containing length of subsequents
            count2=1
            for i in range(len(octs)-1):
                if(octs[i]==2 and octs[i+1]==2):
                    count2+=1
                elif(octs[i]==2 and octs[i+1]!=2):
                    list2.append(count2)
                    count2=1

            list_2=[]      #list containing length of subsequents
            count_2=1
            for i in range(len(octs)-1):
                if(octs[i]==-2 and octs[i+1]==-2):
                    count_2+=1
                elif(octs[i]==-2 and octs[i+1]!=-2):
                    list_2.append(count_2)
                    count_2=1

            list3=[]      #list containing length of subsequents
            count3=1
            for i in range(len(octs)-1):
                if(octs[i]==3 and octs[i+1]==3):
                    count3+=1
                elif(octs[i]==3 and octs[i+1]!=3):
                    list3.append(count3)
                    count3=1

            list_3=[]      #list containing length of subsequents
            count_3=1
            for i in range(len(octs)-1):
                if(octs[i]==-3 and octs[i+1]==-3):
                    count_3+=1
                elif(octs[i]==-3 and octs[i+1]!=-3):
                    list_3.append(count_3)
                    count_3=1

            list4=[]      #list containing length of subsequents
            count4=1
            for i in range(len(octs)-1):
                if(octs[i]==4 and octs[i+1]==4):
                    count4+=1
                elif(octs[i]==4 and octs[i+1]!=4):
                    list4.append(count4)
                    count4=1

            list_4=[]      #list containing length of subsequents
            count_4=1
            for i in range(len(octs)-1):
                if(octs[i]==-4 and octs[i+1]==-4):
                    count_4+=1
                elif(octs[i]==-4 and octs[i+1]!=-4):
                    list_4.append(count_4)
                    count_4=1

            # here we have printed the longest subsequence of octants
            sheet['AT4'] = max(list1)
            sheet['AT5'] = max(list_1)
            sheet['AT6'] = max(list2)
            sheet['AT7'] = max(list_2)
            sheet['AT8'] = max(list3)
            sheet['AT9'] = max(list_3)
            sheet['AT10'] = max(list4)
            sheet['AT11'] = max(list_4)

            # here we have printed the howmany times maximum subsequence have occur
            sheet['AU4'] = list1.count(max(list1))
            sheet['AU5'] = list_1.count(max(list_1))
            sheet['AU6'] = list2.count(max(list2))
            sheet['AU7'] = list_2.count(max(list_2))
            sheet['AU8'] = list3.count(max(list3))
            sheet['AU9'] = list_3.count(max(list_3))
            sheet['AU10'] = list4.count(max(list4))
            sheet['AU11'] = list_4.count(max(list_4))

            for i in range(4,12):
                for j in range(46,48):
                    sheet.cell(row=i,column=j).border = border


                #################################################################################################################
            #################################################################################################################

            for i in range(8):
                sheet.cell(row=3,column=23+i).value = "rank of "+str(new_octant[i])
                sheet.cell(row=3,column=23+i).border = border


            dictionary_ct={1:ctpos1,-1:ctneg1,2:ctpos2,-2:ctneg2,3:ctpos3,-3:ctneg3,4:ctpos4,-4:ctneg4}
            dictionary_ct=dict(sorted(dictionary_ct.items(), key=lambda item:item[1]))
            dictionary_ct=list(dictionary_ct.items())

            for i in range(8):                        # code to print rank 1 in overall count
                if(dictionary_ct[i][0]==1):
                    sheet.cell(row=4,column=23).value=8-i
                    sheet.cell(row=4,column=23).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=23).fill = pattern
                elif(dictionary_ct[i][0]==-1):
                    sheet.cell(row=4,column=24).value=8-i
                    sheet.cell(row=4,column=24).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=24).fill = pattern
                elif(dictionary_ct[i][0]==2):
                    sheet.cell(row=4,column=25).value=8-i
                    sheet.cell(row=4,column=25).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=25).fill = pattern
                elif(dictionary_ct[i][0]==-2):
                    sheet.cell(row=4,column=26).value=8-i
                    sheet.cell(row=4,column=26).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=26).fill = pattern
                elif(dictionary_ct[i][0]==3):
                    sheet.cell(row=4,column=27).value=8-i
                    sheet.cell(row=4,column=27).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=27).fill = pattern
                elif(dictionary_ct[i][0]==-3):
                    sheet.cell(row=4,column=28).value=8-i
                    sheet.cell(row=4,column=28).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=28).fill = pattern
                elif(dictionary_ct[i][0]==4):
                    sheet.cell(row=4,column=29).value=8-i
                    sheet.cell(row=4,column=29).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=29).fill = pattern
                elif(dictionary_ct[i][0]==-4):
                    sheet.cell(row=4,column=30).value=8-i
                    sheet.cell(row=4,column=30).border = border
                    if((8-i)==1):
                        sheet.cell(row=4, column=30).fill = pattern



            sheet['AE3'] = "Rank1 Octant ID"
            sheet['AE3'].border = border
            sheet['AF3'] = "Rank1 Octant Name"
            sheet['AF3'].border = border

            octant_id = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
            sheet['AE4']=dictionary_ct[7][0]
            sheet['AE4'].border = border
            sheet['AF4']=octant_id[str(dictionary_ct[7][0])]
            sheet['AF4'].border = border

            rank1=[]    # this list stores the number of rank 1 octants

            for i in range(p):   # made a dictionary to store the count value of each octants
                mod_dictionary_ct={1:A[i].count(1), -1:A[i].count(-1),2:A[i].count(2),-2:A[i].count(-2),3:A[i].count(3),-3:A[i].count(-3),4:A[i].count(4),-4:A[i].count(-4)}
                mod_dictionary_ct=dict(sorted(mod_dictionary_ct.items(),key=lambda item:item[1]))     # sorted the values in increasing order
                mod_dictionary_ct=list(mod_dictionary_ct.items())    # and make a list of that dictionary

                for j in range(8):                  # code to print the rank of octants in mod values
                    if(mod_dictionary_ct[j][0]==1):
                        sheet.cell(row=5+i,column=23).value=8-j
                        sheet.cell(row=5+i,column=23).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=23).fill = pattern
                    elif(mod_dictionary_ct[j][0]==-1):
                        sheet.cell(row=5+i,column=24).value=8-j
                        sheet.cell(row=5+i,column=24).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=24).fill = pattern
                    elif(mod_dictionary_ct[j][0]==2):
                        sheet.cell(row=5+i,column=25).value=8-j
                        sheet.cell(row=5+i,column=25).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=25).fill = pattern
                    elif(mod_dictionary_ct[j][0]==-2):
                        sheet.cell(row=5+i,column=26).value=8-j
                        sheet.cell(row=5+i,column=26).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=26).fill = pattern
                    elif(mod_dictionary_ct[j][0]==3):
                        sheet.cell(row=5+i,column=27).value=8-j
                        sheet.cell(row=5+i,column=27).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=27).fill = pattern
                    elif(mod_dictionary_ct[j][0]==-3):
                        sheet.cell(row=5+i,column=28).value=8-j
                        sheet.cell(row=5+i,column=28).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=28).fill = pattern
                    elif(mod_dictionary_ct[j][0]==4):
                        sheet.cell(row=5+i,column=29).value=8-j
                        sheet.cell(row=5+i,column=29).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=29).fill = pattern
                    elif(mod_dictionary_ct[j][0]==-4):
                        sheet.cell(row=5+i,column=30).value=8-j
                        sheet.cell(row=5+i,column=30).border = border
                        if((8-j)==1):
                            sheet.cell(row=5+i, column=30).fill = pattern

                sheet.cell(row=5+i,column=31).value=mod_dictionary_ct[7][0]
                sheet.cell(row=5+i,column=31).border = border
                rank1.append(mod_dictionary_ct[7][0])                   
                sheet.cell(row=5+i,column=32).value=octant_id[str(mod_dictionary_ct[7][0])]
                sheet.cell(row=5+i,column=32).border = border



            # code to print the octant which 1 rank appear most

            sheet.cell(row=6+p,column=29).value="Octant ID"
            sheet.cell(row=6+p,column=29).border = border
            sheet.cell(row=6+p,column=30).value="Octant Name"
            sheet.cell(row=6+p,column=30).border= border
            sheet.cell(row=6+p,column=31).value="Count of Rank 1 Mod Values"
            sheet.cell(row=6+p,column=31).border=border

            for i in range(8):                           # code for rank 1 ID
                sheet.cell(row=7+p+i,column=29).value=str(new_octant[i])
                sheet.cell(row=7+p+i,column=29).border = border
                sheet.cell(row=7+p+i,column=30).value=octant_id[str(new_octant[i])]
                sheet.cell(row=7+p+i,column=30).border = border
                sheet.cell(row=7+p+i,column=31).value=rank1.count(new_octant[i])
                sheet.cell(row=7+p+i,column=31).border = border

            #################################################################################################################
            #################################################################################################################

            time = []
            for i in range(len(l1)):
                x=sheet.cell(row=i+2,column=1).value
                time.append(x)

            row_no1 = []     #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list1=[]      #list containing length of subsequents
            count1=1
            for i in range(len(octs)-1):
                if(octs[i]==1 and octs[i+1]==1):
                    count1+=1
                elif(octs[i]==1 and octs[i+1]!=1):
                    list1.append(count1)
                    row_no1.append(i-count1+1)
                    count1=1

            time_l1 = []        #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list1)-1):
                if(list1[i]==max(list1)):
                    a=row_no1[i]
                    time_l1.append([time[a],time[a+max(list1)-1]])

            row_no_1 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list_1=[]      #list containing length of subsequents
            count_1=1
            for i in range(len(octs)-1):
                if(octs[i]==-1 and octs[i+1]==-1):
                    count_1+=1
                elif(octs[i]==-1 and octs[i+1]!=-1):
                    list_1.append(count_1)
                    row_no_1.append(i-count_1+1)
                    count_1=1

            time_l_1 = []           #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list_1)-1):
                if(list_1[i]==max(list_1)):
                    a=row_no_1[i]
                    time_l_1.append([time[a],time[a+max(list_1)-1]])


            row_no2 = []            #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list2=[]      #list containing length of subsequents
            count2=1
            for i in range(len(octs)-1):
                if(octs[i]==2 and octs[i+1]==2):
                    count2+=1
                elif(octs[i]==2 and octs[i+1]!=2):
                    list2.append(count2)
                    row_no2.append(i-count2+1)
                    count2=1

            time_l2 = []          #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list2)-1):
                if(list2[i]==max(list2)):
                    a=row_no2[i]
                    time_l2.append([time[a],time[a+max(list2)-1]])

            row_no_2 = []       #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list_2=[]      #list containing length of subsequents
            count_2=1
            for i in range(len(octs)-1):
                if(octs[i]==-2 and octs[i+1]==-2):
                    count_2+=1
                elif(octs[i]==-2 and octs[i+1]!=-2):
                    list_2.append(count_2)
                    row_no_2.append(i-count_2+1)
                    count_2=1

            time_l_2 = []         #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list_2)-1):
                if(list_2[i]==max(list_2)):
                    a=row_no_2[i]
                    time_l_2.append([time[a],time[a+max(list_2)-1]])

            row_no3 = []           #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list3=[]      #list containing length of subsequents
            count3=1
            for i in range(len(octs)-1):
                if(octs[i]==3 and octs[i+1]==3):
                    count3+=1
                elif(octs[i]==3 and octs[i+1]!=3):
                    list3.append(count3)
                    row_no3.append(i-count3+1)
                    count3=1

            time_l3 = []           #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list3)-1):
                if(list3[i]==max(list3)):
                    a=row_no3[i]
                    time_l3.append([time[a],time[a+max(list3)-1]])

            row_no_3 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list_3=[]      #list containing length of subsequents
            count_3=1
            for i in range(len(octs)-1):
                if(octs[i]==-3 and octs[i+1]==-3):
                    count_3+=1
                elif(octs[i]==-3 and octs[i+1]!=-3):
                    list_3.append(count_3)
                    row_no_3.append(i-count_3+1)
                    count_3=1

            time_l_3 = []             #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list_3)-1):
                if(list_3[i]==max(list_3)):
                    a=row_no_3[i]
                    time_l_3.append([time[a],time[a+max(list_3)-1]])

            row_no4 = []          #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list4=[]      #list containing length of subsequents
            count4=1
            for i in range(len(octs)-1):
                if(octs[i]==4 and octs[i+1]==4):
                    count4+=1
                elif(octs[i]==4 and octs[i+1]!=4):
                    list4.append(count4)
                    row_no4.append(i-count4+1)
                    count4=1

            time_l4 = []            #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list4)-1):
                if(list4[i]==max(list4)):
                    a=row_no4[i]
                    time_l4.append([time[a],time[a+max(list4)-1]])

            row_no_4 = []        #we have created this list to store the row number of the row form which a subsequence is starting. this row number is the index of list of time.
            list_4=[]      #list containing length of subsequents
            count_4=1
            for i in range(len(octs)-1):
                if(octs[i]==-4 and octs[i+1]==-4):
                    count_4+=1
                elif(octs[i]==-4 and octs[i+1]!=-4):
                    list_4.append(count_4)
                    row_no_4.append(i-count_4+1)
                    count_4=1

            time_l_4 = []              #in this list we are collecting the range of time of a subsequence.
            for i in range(len(list_4)-1):
                if(list_4[i]==max(list_4)):
                    a=row_no_4[i]
                    time_l_4.append([time[a],time[a+max(list_4)-1]])


            sheet.cell(row=1,column=49).value = "Longest Subsquence Length with Range"
            sheet.cell(row=3,column=49).value = "Octant ##"
            sheet.cell(row=3,column=50).value = "Longest Subsequence Length"
            sheet.cell(row=3,column=51).value = "Count"

            octant=[1,-1,2,-2,3,-3,4,-4]
            i=4                                                  # the whole code is to print the time range of longest subsequence for octants
            for oct in octant:
                sheet.cell(row=i+1,column=49).value="Time"
                sheet.cell(row=i+1,column=50).value="From"
                sheet.cell(row=i+1,column=51).value="To"
                if oct==1:
                    sheet.cell(row=i,column=49).value="+1"
                    sheet.cell(row=i,column=50).value=max(list1)
                    sheet.cell(row=i,column=51).value=list1.count(max(list1))

                    l=0
                    while l<len(time_l1):
                        sheet.cell(row=i+2+l,column=50).value=time_l1[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l1[l][1]
                        l+=1

                elif oct==-1:
                    sheet.cell(row=i,column=49).value="-1"
                    sheet.cell(row=i,column=50).value=max(list_1)
                    sheet.cell(row=i,column=51).value=list_1.count(max(list_1))

                    l=0
                    while l<len(time_l_1):
                        sheet.cell(row=i+2+l,column=50).value=time_l_1[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l_1[l][1]
                        l+=1

                elif oct==2:
                    sheet.cell(row=i,column=49).value="+2"
                    sheet.cell(row=i,column=50).value=max(list2)
                    sheet.cell(row=i,column=51).value=list2.count(max(list2))

                    l=0
                    while l<len(time_l2):
                        sheet.cell(row=i+2+l,column=50).value=time_l2[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l2[l][1]
                        l+=1

                elif oct==-2:
                    sheet.cell(row=i,column=49).value="-2"
                    sheet.cell(row=i,column=50).value=max(list_2)
                    sheet.cell(row=i,column=51).value=list_2.count(max(list_2))

                    l=0
                    while l<len(time_l_2):
                        sheet.cell(row=i+2+l,column=50).value=time_l_2[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l_2[l][1]
                        l+=1

                elif oct==3:
                    sheet.cell(row=i,column=49).value="+3"
                    sheet.cell(row=i,column=50).value=max(list3)
                    sheet.cell(row=i,column=51).value=list3.count(max(list3))

                    l=0
                    while l<len(time_l3):
                        sheet.cell(row=i+2+l,column=50).value=time_l3[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l3[l][1]
                        l+=1

                elif oct==-3:
                    sheet.cell(row=i,column=49).value="-3"
                    sheet.cell(row=i,column=50).value=max(list_3)
                    sheet.cell(row=i,column=51).value=list_3.count(max(list_3))

                    l=0
                    while l<len(time_l_3):
                        sheet.cell(row=i+2+l,column=50).value=time_l_3[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l_3[l][1]
                        l+=1

                elif oct==4:
                    sheet.cell(row=i,column=49).value="+4"
                    sheet.cell(row=i,column=50).value=max(list4)
                    sheet.cell(row=i,column=51).value=list4.count(max(list4))

                    l=0
                    while l<len(time_l4):
                        sheet.cell(row=i+2+l,column=50).value=time_l4[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l4[l][1]
                        l+=1

                elif oct==-4:
                    sheet.cell(row=i,column=49).value="-4"
                    sheet.cell(row=i,column=50).value=max(list_4)
                    sheet.cell(row=i,column=51).value=list_4.count(max(list_4))

                    l=0
                    while l<len(time_l_4):
                        sheet.cell(row=i+2+l,column=50).value=time_l_4[l][0]
                        sheet.cell(row=i+2+l,column=51).value=time_l_4[l][1]
                        l+=1

                i+=l+2


            #####################################################

            for k in range(3,30):
                for l in range(49,52):
                    sheet.cell(row=k, column=l).border = border



            for column in ["E","F","G","H","I","J","W","X","Y","Z","AB","AA","AC"]:
                sheet.column_dimensions[column].width = 15                           # have set the width of the some column

            for column in ["AD","AE","AF","AI","AS","AT","AX"]:
                sheet.column_dimensions[column].width = 30

            for column in ["O","P","Q","R","S","T","U","V","AJ","AK","AL","AM","AN","AO","AP","AQ","AU","AY"]:
                sheet.column_dimensions[column].width = 6

            for column in ["AW"]:
                sheet.column_dimensions[column].width = 38
            for column in ["N"]:
                sheet.column_dimensions[column].width = 22

            time2=datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            wb.save('output/'+str(file.name[:-5]) + '_mod' + '(' + str(mod)+ ')_' + str(time2) + '.xlsx')


        st.info("Completed !!")
  
end_time = datetime.now()

print("Duraion of Program Execution: ",end_time-start_time)
